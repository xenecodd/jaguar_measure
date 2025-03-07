import os
import threading
import time
import numpy as np
import open3d as o3d
import pandas as pd
import mysql.connector
import matplotlib

# backend kontrolü: use_agg True ise non-interaktif, Agg backend; False ise interaktif, TkAgg
USE_AGG = True

if USE_AGG:
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    # Agg modda GUI gösterimlerini kapatıyoruz
    plt.show = lambda: None
else:
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt

from tkinter.messagebox import showerror
from openpyxl.styles import PatternFill
from robot_control import send_command
from Measure.Scripts import *
from mecheye_trigger import TriggerWithExternalDeviceAndFixedRate, robot
from points import *
# robot.SetDO(7,0)
# time.sleep(2)
class JaguarScanner:
    def __init__(self,vel_mul, use_agg: bool = USE_AGG):
        # use_agg True ise Agg mod (non-interaktif, thread’li hesaplama); False ise interaktif, threadsiz hesaplama
        self.use_agg = use_agg
        self.mech_eye = TriggerWithExternalDeviceAndFixedRate(vel_mul)
        self.pcd = o3d.geometry.PointCloud()
        self.results = []
        self.excel_threads = []
        self.points = left_of_robot_points + left_small + back_points + right_of_robot_points + right_small 

        self.file_path = "/home/eypan/Documents/down_jaguar/point_index.txt"
        self.vel_mul = vel_mul
        self.ignored_points = []
        self.same_object = False
        self.same_place_index = None
        self.drop_object = True
        self.pick = True
        self.save_point_clouds = False
        self.save_to_db = False
        self.range_ = 70

        # Toleranslar: (hedef, tolerans)
        self.tolerances = {
            "Feature1 (102.1)": (102.1, 2.0),
            "Feature2 (25mm/2)": (12.5, 0.5),
            "Feature3 (23.1)": (23.1, 1),
            "Feature4 (25mm/2)": (12.5, 0.5),
            "Feature5 (L40)": (40.0, 1),
            "Feature6 (L248)": (248.0, 2.0),
            "Feature7 (L42)": (42.0, 1.5),
            "Feature8 (L79.73)": (79.73, 1.5),
            "Feature9 (R1-50)": (50, 1.5),
            "Feature10 (R2-35)": (35, 1.5),
            "Feature11 (3mm)": (0, 3.0),
            "Feature12 (88.6)": (88.6, 1.5),
            "Feature13 (10.6)": (10.6, 1.0),
            "Feature14 (81.5)": (81.5, 1.5),
            "Feature15 (L23.4)": (23.4, 1.0),
            "Feature16 (L17.2)": (17.2, 1.0),
            "Feature17 (2C)": (0.0, 2.0)
        }

    # ----- Yardımcı Fonksiyonlar -----
    @staticmethod
    def rotate_point_cloud(points: np.ndarray, angle_degrees: float, axis: str) -> np.ndarray:
        angle_radians = np.radians(angle_degrees)
        if axis == "z":
            rotation_matrix = np.array([
                [np.cos(angle_radians), -np.sin(angle_radians), 0],
                [np.sin(angle_radians),  np.cos(angle_radians), 0],
                [0, 0, 1]
            ])
        elif axis == "x":
            rotation_matrix = np.array([
                [1, 0, 0],
                [0, np.cos(angle_radians), -np.sin(angle_radians)],
                [0, np.sin(angle_radians),  np.cos(angle_radians)]
            ])
        elif axis == "y":
            rotation_matrix = np.array([
                [np.cos(angle_radians), 0, np.sin(angle_radians)],
                [0, 1, 0],
                [-np.sin(angle_radians), 0, np.cos(angle_radians)]
            ])
        else:
            raise ValueError("Desteklenmeyen eksen. 'x', 'y' veya 'z' girin.")
        return points @ rotation_matrix.T

    @staticmethod
    def to_origin(points: np.ndarray) -> np.ndarray:
        min_x, min_y = np.min(points[:, 0]), np.min(points[:, 1])
        points[:, 0] -= min_x
        points[:, 1] -= min_y
        return points

    @staticmethod
    def remove_gripper_points(points: np.ndarray) -> np.ndarray:
        min_x = np.min(points[:, 0])
        y_candidates = points[:, 1][points[:, 0] < min_x + 23]
        y_min, y_max = np.min(y_candidates), np.max(y_candidates)
        return points[(points[:, 1] < y_min) | (points[:, 1] > y_max)]

    def get_next_valid_index(self, current_index: int, ignored: list, total_points: int) -> int:
        next_index = (current_index + 1) % total_points
        while ignored and next_index in ignored:
            next_index = (next_index + 1) % total_points
        return next_index

    def read_current_point_index(self) -> int:
        if os.path.exists(self.file_path):
            with open(self.file_path, 'r') as file:
                return int(file.read().strip())
        return 0

    def write_current_point_index(self, index: int):
        with open(self.file_path, 'w') as file:
            file.write(str(index))

    def pick_object(self, point: dict,soft_point,use_back_transit):
        robot.MoveCart(soft_point, 0, 0, vel=self.vel_mul*100)
        if point["p"][0]>0:
            transit_vel=20
        if use_back_transit:
            print("BACK TRANSİTTT")
            transit_vel = 50
            robot.MoveCart(back_transit_point, 0, 0, vel=self.vel_mul * transit_vel)
        robot.MoveCart(point["p_up"], 0, 0, vel=self.vel_mul*transit_vel)
        robot.MoveL(point["p"], 0, 0, vel=self.vel_mul*50)
        robot.WaitMs(500)
        robot.SetDO(7, 1)
        robot.WaitMs(500)
        robot.MoveL(point["p_up"], 0, 0, vel=self.vel_mul*transit_vel)
        if use_back_transit:
            robot.MoveCart(back_transit_point, 0, 0, vel=self.vel_mul * transit_vel)
        robot.MoveCart(soft_point, 0, 0, vel=self.vel_mul*transit_vel)

    def calculate_tolerance_distance(self, value: float, target: float, tolerance: float) -> float:
        return np.abs(value - target)

    def get_gradient_color(self, distance: float, tolerance: float) -> str:
        if distance > tolerance:
            return "FF0000"
        ratio = distance / tolerance
        red   = int(255 * (1 - ratio))
        green = int(255 * (1 - ratio))
        blue  = 255
        return f"{red:02X}{green:02X}{blue:02X}"

    # ----- Ölçüm İşleme Fonksiyonları -----
    # Tarama verileri alındıktan sonra hesaplamalar, use_agg True ise ayrı thread’lerde paralel, 
    # False ise sıralı (threadsiz) çalıştırılır.

    def smol_calc(self, small_data: np.ndarray):
        small = self.rotate_point_cloud(small_data, -90, "z")
        small = small[small[:, 2] > np.min(small[:, 2]) + 37]
        small = small[small[:, 0] < np.min(small[:, 0]) + 50]
        small = self.to_origin(small)

        if self.save_point_clouds:
            self.pcd.points = o3d.utility.Vector3dVector(small)
            o3d.io.write_point_cloud("/home/eypan/Documents/down_jaguar/jaguar_measure/small.ply", self.pcd)

        z_max = np.max(small[:, 2])
        circle_fitter = CircleFitter(small)
        # Eğer plot yapılması gerekiyorsa, CircleFitter içerisindeki fit_circles_and_plot() çağrılır.
        # Agg modda grafik gösterimi devre dışı bırakıldığı için plt.show no-op olacak.
        _, z_center_small, radius_small = circle_fitter.fit_circles_and_plot(
            find_second_circle=False, val_x=0.18, val_z=0.2, delta_z=25
        )
        s_datum = circle_fitter.get_datum()
        self.dist_3mm_s, _ = circle_fitter.get_distance(second_crc=False, z_distance_to_datum=23.1)
        self.feature_3 = z_center_small - s_datum
        self.radius_small = radius_small
        self.z_center_small = z_center_small

    def hor_calc(self, horizontal_data: np.ndarray, horizontal2_data: np.ndarray):
        horizontal2_data[:, 2] -= 70
        horizontal2_data[:, 0] -= 50

        if self.save_point_clouds:
            self.pcd.points = o3d.utility.Vector3dVector(horizontal2_data)
            o3d.io.write_point_cloud("/home/eypan/Documents/down_jaguar/jaguar_measure/horizontal2.ply", self.pcd)
            self.pcd.points = o3d.utility.Vector3dVector(horizontal_data)
            o3d.io.write_point_cloud("/home/eypan/Documents/down_jaguar/jaguar_measure/horizontal_pre.ply", self.pcd)

        diff = np.abs(np.max(horizontal_data[:, 0]) - np.min(horizontal2_data[:, 0]))
        print("diff", diff)
        datum_horizontal = np.max(horizontal_data[:, 0]) - diff

        y_values = np.linspace(np.min(horizontal_data[:, 1]), np.max(horizontal_data[:, 1]), num=100)
        line_points = np.array([[datum_horizontal, y, np.min(horizontal_data[:, 2])] for y in y_values])
        augmented_pc = np.vstack((horizontal_data, line_points))
        horizontal = self.rotate_point_cloud(augmented_pc, 90, "z")
        horizontal = self.to_origin(horizontal)
        self.horizontal = horizontal

        if self.save_point_clouds:
            self.pcd.points = o3d.utility.Vector3dVector(horizontal)
            o3d.io.write_point_cloud("/home/eypan/Documents/down_jaguar/jaguar_measure/horizontal_post.ply", self.pcd)

        self.circle_fitter = CircleFitter(horizontal)
        _, circle2 = self.circle_fitter.fit_circles_and_plot()
        self.dist_3mm_h = self.circle_fitter.get_distance()[0]
        self.l_40 = get_40(horizontal)
        self.height = np.max(self.horizontal[:, 1]) - self.circle_fitter.get_datum()
        self.feature_1 = circle2[1]
        self.feature_2 = circle2[2]

    def process_vertical_measurement(self, vertical_data: np.ndarray) -> dict:
        vertical = self.remove_gripper_points(vertical_data)
        vertical = self.rotate_point_cloud(vertical, 180, "z")

        if self.save_point_clouds:
            pcd = o3d.geometry.PointCloud()
            pcd.points = o3d.utility.Vector3dVector(vertical)
            o3d.io.write_point_cloud("/home/eypan/Documents/down_jaguar/jaguar_measure/vertical.ply", pcd)
        vertical_copy = self.to_origin(vertical.copy())
        l_17_2, ok_17_2 = horn_diff(vertical_copy)
        l_23_4, ok_23_4 = horn_diff(vertical_copy, 240, 280)

        current_index = self.get_next_valid_index(self.read_current_point_index(), self.ignored_points, len(self.points))
        self.write_current_point_index(current_index)
        if self.same_object and self.cycle < self.range_:
            point = self.points[current_index]
            robot.MoveCart(point["p_up"], 0, 0, vel=self.vel_mul*80)
            robot.MoveL(point["p"], 0, 0, vel=self.vel_mul*50)
            robot.WaitMs(500)
            robot.SetDO(7, 0)
            robot.WaitMs(500)
            robot.MoveL(point["p_up"], 0, 0, vel=self.vel_mul*100)
        elif self.drop_object:
            print("DEVAM")
            robot.MoveCart(self.trash, 0, 0, vel=self.vel_mul*80 )
            robot.SetDO(7, 0)
        if self.cycle == self.range_ - 1:
            self.write_current_point_index(0)

        B = self.circle_fitter.get_B()
        print("B:", B)
        b_trans_val = np.max(vertical[:, 1]) - np.max(self.horizontal[:, 2])
        b_vertical = B + b_trans_val
        print("b_vertical:", b_vertical)
        _, _, r1, l_79_73 = slope(vertical, b_vertical)
        _, _, r2, _ = slope(vertical, y_divisor=0.11, crc_l=27)
        l_42 = np.max(vertical[:, 1]) - b_vertical
        l_248 = arm_horn_lengths(vertical, b_vertical)
        mean_3mm = np.mean([self.dist_3mm_h, self.dist_3mm_s])
        l_88_6 = self.feature_1 - self.feature_2
        l_81_5 = filter_and_visualize_projection_with_ply(self.horizontal)

        return {
            "l_17_2": l_17_2,
            "l_23_4": l_23_4,
            "l_42": l_42,
            "l_79_73": l_79_73,
            "l_248": l_248,
            "r1": (r1 - self.feature_2),
            "r2": (r2 + self.feature_2),
            "feature_1": self.feature_1,
            "mean_3mm": mean_3mm,
            "l_88_6": l_88_6,
            "l_81_5": l_81_5,
            "ok_17_2": ok_17_2
        }

    def combine_results(self, vertical_results: dict) -> dict:
        return {
            "Feature1 (102.1)": vertical_results.get("feature_1") or self.feature_1,
            "Feature2 (25mm/2)": self.feature_2,
            "Feature3 (23.1)": self.feature_3,
            "Feature4 (25mm/2)": self.radius_small,
            "Feature5 (L40)": self.l_40,
            "Feature6 (L248)": vertical_results.get("l_248"),
            "Feature7 (L42)": vertical_results.get("l_42"),
            "Feature8 (L79.73)": vertical_results.get("l_79_73"),
            "Feature9 (R1-50)": vertical_results.get("r1"),
            "Feature10 (R2-35)": vertical_results.get("r2"),
            "Feature11 (3mm)": vertical_results.get("mean_3mm"),
            "Feature12 (88.6)": vertical_results.get("l_88_6"),
            "Feature13 (10.6)": self.feature_3 - self.radius_small,
            "Feature14 (81.5)": vertical_results.get("l_81_5"),
            "Feature15 (L23.4)": vertical_results.get("l_23_4"),
            "Feature16 (L17.2)": vertical_results.get("l_17_2"),
            "Feature17 (2C)": vertical_results.get("ok_17_2")
        }

    def write_excel_to_db(self, excel_path):
        # Excel dosyasını oku
        df = pd.read_excel(excel_path, sheet_name="ScanResults")
        
        # MariaDB'ye bağlan
        conn = mysql.connector.connect(
            host="192.168.1.180",
            user="cobot_dbuser",
            password="um6vv$7*sJ@5Q*",
            database="cobot"
        )
        cursor = conn.cursor()

        # Tabloyu oluştur (eğer yoksa)
        create_table_query = """
        CREATE TABLE IF NOT EXISTS scan_results (
            id INT AUTO_INCREMENT PRIMARY KEY,
            iteration INT,
            feature VARCHAR(255),
            value DOUBLE
        )
        """
        cursor.execute(create_table_query)
        conn.commit()

        # Excel'den okunan verileri tabloya ekle;
        # Eğer value sayısal değilse, None olarak eklenir.
        insert_query = "INSERT INTO scan_results (iteration, feature, value) VALUES (%s, %s, %s)"
        for index, row in df.iterrows():
            try:
                value_float = float(row['Value'])
            except ValueError:
                value_float = None
            data = (int(row['Iteration']), row['Feature'], value_float)
            cursor.execute(insert_query, data)
        
        conn.commit()
        cursor.close()
        conn.close()
        print("Excel verileri başarıyla veritabanına yazıldı.")

    def save_to_excel(self):
        rows = []
        for iteration, result in enumerate(self.results, start=1):
            for feature, value in result.items():
                rows.append({"Iteration": iteration, "Feature": feature, "Value": value})
        results_df = pd.DataFrame(rows)
        output_file = "/home/eypan/Documents/down_jaguar/jaguar_measure/scan_results.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, index=False, sheet_name="ScanResults")
                workbook = writer.book
                worksheet = writer.sheets["ScanResults"]
                for row_index, row in enumerate(results_df.itertuples(index=False), start=2):
                    iteration = row.Iteration
                    feature = row.Feature
                    value = row.Value
                    iteration_color = "FCE4D6" if iteration % 2 == 0 else "D9EAD3"
                    fill = PatternFill(start_color=iteration_color, end_color=iteration_color, fill_type="solid")
                    for col_index in range(1, len(results_df.columns) + 1):
                        worksheet.cell(row=row_index, column=col_index).fill = fill
                    if feature in self.tolerances:
                        target, tolerance = self.tolerances[feature]
                        value_cell = worksheet.cell(row=row_index, column=3)
                        if target - tolerance <= value <= target + tolerance:
                            value_cell.fill = PatternFill(start_color="00B050", fill_type="solid")
                        else:
                            value_cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                        distance = self.calculate_tolerance_distance(value, target, tolerance)
                        gradient_color = self.get_gradient_color(distance, tolerance)
                        distance_cell = worksheet.cell(row=row_index, column=4)
                        distance_cell.value = tolerance - distance
                        distance_cell.fill = PatternFill(start_color=gradient_color, fill_type="solid")
            if self.save_to_db:            
                self.write_excel_to_db(output_file)
        except FileNotFoundError:
            print(f"Error: Dizin bulunamadı: {output_file}")
        print("Excel'e kaydetme işlemi tamamlandı.")

    # ----- Tarama Döngüsü -----
    # Robot taramaları sıralı yapılır. Alınan tarama verileriyle,
    # use_agg=True ise small ve horizontal hesaplamaları ayrı thread’lerde paralel çalışır,
    # aksi halde sıralı (threadsiz) çalıştırılır.
    
    def run_scan_cycle(self):
        ROBOT_POSITIONS = {
            'scrc': [-450, -130, 470, 82.80, 89.93, -7.30],
            'h_2': [-375, 100, 545, -90, -90, 180],
            'p_91': [-335, 100, 350, -90.00, -0.0005, 90.00],
            'notOK': [-621, 325, 511, 117, -83, -148]
        }
        MAX_RETRIES = 2  # Örnek sayı

        for self.cycle in range(self.range_):
            start_time = time.time()
            current_results = {}
            current_index = self.read_current_point_index()
            try:
                send_command({"cmd": 107, "data": {"content": "ResetAllError()"}})

                if self.pick:
                    soft_point = p90 if current_index <= len(left_of_robot_points + left_small + back_points) else p91
                    self.trash = p91_trash if soft_point == p91 else p90_trash
                    print("CURRENT İNDEX",current_index)
                    use_back_transit = True if current_index >= len(left_of_robot_points + left_small) and current_index <= len(left_of_robot_points + left_small + back_points) else False
                    print("Use back transit:", use_back_transit)
                    point = self.points[self.same_place_index] if isinstance(self.same_place_index, int) else self.points[current_index]
                    self.pick_object(point,soft_point,use_back_transit)


                robot.MoveCart(ROBOT_POSITIONS['scrc'], 0, 0, vel=self.vel_mul*100)

                # --- Small hesaplama ---
                small_data = self.mech_eye.main(lua_name="small.lua", scan_line_count=1500)
                if self.use_agg:
                    thread_small = threading.Thread(target=self.smol_calc, args=(small_data,))
                    thread_small.start()
                else:
                    self.smol_calc(small_data)
                    plt.show()

                # --- Horizontal hesaplama ---
                horizontal_data = self.mech_eye.main(lua_name="horizontal.lua", scan_line_count=1500)
                horizontal2_data = self.mech_eye.main(lua_name="horizontal2.lua", scan_line_count=1500)
                if self.use_agg:
                    thread_horizontal = threading.Thread(target=self.hor_calc, args=(horizontal_data, horizontal2_data))
                    thread_horizontal.start()
                else:
                    self.hor_calc(horizontal_data, horizontal2_data)
                    plt.show()

                # --- Vertical tarama ---
                vertical_data = self.mech_eye.main(lua_name="vertical.lua", scan_line_count=2500)
                plt.show()

                # Eğer thread'ler kullanıldıysa, bunların tamamlanmasını bekle
                if self.use_agg:
                    thread_small.join()
                    thread_horizontal.join()

                # Vertical hesaplama, horizontal sonuçlara bağlı olduğundan şimdi yapılır.
                vertical_results = self.process_vertical_measurement(vertical_data)
                current_results = self.combine_results(vertical_results)
                current_results["Processing Time (s)"] = time.time() - start_time
                print("Cycle süresi:", time.time() - start_time)
            except Exception as e:
                print(f"Cycle {self.cycle + 1} hata verdi: {str(e)}")
                current_results = {"Error": str(e)}
            finally:
                self.results.append(current_results)
                # Excel'e kaydetme işlemi; use_agg True ise thread ile, aksi halde doğrudan.
                if self.use_agg:
                    t_excel = threading.Thread(target=self.save_to_excel)
                    t_excel.start()
                    self.excel_threads.append(t_excel)
                else:
                    self.save_to_excel()
        # İsteğe bağlı: Tüm Excel thread'lerinin bitmesini bekleyebilirsiniz.
        # for t in self.excel_threads:
        #     t.join()


if __name__ == "__main__":
    # use_agg=True: Agg mod, thread'li hesaplama (GUI kapalı, plt.show no-op)
    # use_agg=False: interaktif mod, thread kullanılmadan ana thread'de hesaplama
    
    scanner = JaguarScanner(vel_mul=1) # vel_mul==>> Velocity Multiplier
    scanner.run_scan_cycle()
