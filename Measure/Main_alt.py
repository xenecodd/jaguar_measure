import os
from tkinter.messagebox import showerror
from MecheyePackage.robot_control import send_command
from Scripts import *
from MecheyePackage import TriggerWithExternalDeviceAndFixedRate, robot
import open3d as o3d
import numpy as np
import matplotlib
matplotlib.use('Agg')
import time
import pandas as pd
from openpyxl.styles import PatternFill

class JaguarScanner:
    def __init__(self):
        self.mech_eye = TriggerWithExternalDeviceAndFixedRate()
        self.rotated_pcd = o3d.geometry.PointCloud()
        self.results = []
        self.points = [
            {"p_up": [-747.6768798828125, -36.60947036743164, 224.6449432373047, 179.9124450683593, 0.01771553605794907, 90.00694274902342], 
             "p": [-747.6768188476562, -36.60196685791015, 139.29833984375, 179.9137420654297, 0.016725072637200356, 88.00242614746092]},
            {"p_up": [-629, -36.60947036743164, 224.6449432373047, 179.9124450683593, 0.01771553605794907, 88.00694274902342], 
             "p": [-629, -36.60947036743164, 139.29833984375, 179.9124450683593, 0.01771553605794907, 88.5]},
            {"p_up": [-519, -58.2884407043457, 220.5605010986328, 179.9501495361328, 0.01032531540840864, 88.00717163085938], 
             "p": [-519, -58.29881286621093, 135.5182800292968, 179.9518737792968, 0.010120976716279985, 88.0091323852539]},
            {"p_up": [-398.1421813964843, -58.29881286621093, 220.5605010986328, 179.9518737792968, 0.010120976716279985, 88.0091323852539], 
             "p": [-398.1421813964843, -58.2884407043457, 135.5182800292968, 179.9501495361328, 0.01032531540840864, 88.00717163085938]}
        ]
        self.file_path = "/home/eypan/Documents/alt_jaguar/point_index.txt"

        self.ignored_points = [2]
        self.same_object = False
        self.same_place_index = False
        self.range_ = 3

        self.tolerances = {
            "Feature1 (102.1)": (102.1, 2.0),
            "Feature2 (25mm/2)": (12.5, 0.5),
            "Feature3 (23.1)": (23.1, 1),
            "Feature4 (25mm/2)": (12.5, 0.5),
            "Feature5 (L40)": (40.0, 1.5),
            "Feature6 (L248)": (248.0, 2.0),
            "Feature7 (L42)": (42.0, 1.5),
            "Feature8 (L79.73)": (79.73, 1.5),
            "Feature9 (R1-50)": ((50), 1.5),
            "Feature10 (R2-35)": ((35), 1.5),
            "Feature11 (3mm)": (0, 1.5),
            "Feature12 (88.6)": (88.6, 1.5),
            "Feature13 (10.6)": (10.6, 1.0),
            "Feature14 (81.5)": (81.5, 1.5),
            "Feature15 (L23.4)": (23.4, 1.0),
            "Feature16 (L17.2)": (17.2, 1.0),
        }


    # Yardımcı fonksiyonlar
    @staticmethod
    def rotate_point_cloud(points, angle_degrees, rot):
        """
        Bir nokta bulutunu Z ekseni etrafında belirli bir açıyla döndür.
        :param points: Nx3 boyutunda numpy dizisi (nokta bulutu)
        :param angle_degrees: Döndürme açısı (derece cinsinden)
        :return: Döndürülmüş nokta bulutu (Nx3 numpy dizisi)
        """
        # Dereceden radyana çevir
        angle_radians = np.radians(angle_degrees)
        
        # Z ekseni etrafında dönüşüm matrisi
        if rot == "z":
            
            rotation_matrix = np.array([
            [np.cos(angle_radians), -np.sin(angle_radians), 0],
            [np.sin(angle_radians),  np.cos(angle_radians), 0],
            [0, 0, 1]
        ])
        elif rot == "x":

            rotation_matrix = np.array([
            [1, 0, 0],
            [0, np.cos(angle_radians), -np.sin(angle_radians)],
            [0, np.sin(angle_radians), np.cos(angle_radians)]
        ])
        elif rot == "y":
                
            rotation_matrix = np.array([
            [np.cos(angle_radians), 0, np.sin(angle_radians)],
            [0, 1, 0],
            [-np.sin(angle_radians), 0, np.cos(angle_radians)]
        ])
        
        # Noktaları döndür
        rotated_points = points @ rotation_matrix.T
        return rotated_points

    def get_next_valid_index(self, current_index, ignored_points, total_points):
        """Bir sonraki geçerli nokta indexini hesaplar."""
        next_index = (current_index + 1) % total_points
        while (ignored_points and next_index in (ignored_points)):  # Eğer bir sonraki nokta atlanacaksa bir sonrakine geç
            next_index = (next_index + 1) % total_points
        return next_index

    def pick_object(self, point):
        # Robot hareketleri
        robot.MoveCart(point["p_up"], 0, 0, vel=100)
        robot.MoveL(point["p"], 0, 0, vel=100)
        robot.WaitMs(500)
        robot.SetDO(7, 1)
        robot.WaitMs(500)
        robot.MoveL(point["p_up"], 0, 0, vel=100)

    @staticmethod
    def remove_gripper_points(points):
        """
        Belirtilen Y ekseni aralığında yer alan noktaları bir nokta bulutundan siler.

        Args:
            point_cloud (o3d.geometry.PointCloud): Giriş nokta bulutu.
            y_min (float): Silinecek aralığın alt sınırı (dahil).
            y_max (float): Silinecek aralığın üst sınırı (dahil).

        Returns:
            o3d.geometry.PointCloud: Filtrelenmiş nokta bulutu.
        """
        min_x = np.min(points[:, 0])
        
        y_min, y_max = np.min(points[:,1][points[:,0] < min_x+23]), np.max(points[:,1][points[:,0] < min_x+23])
        # Y ekseni değerine göre filtreleme
        filtered_points = points[(points[:, 1] < y_min) | (points[:, 1] > y_max)]

        return filtered_points

    def process_small(self):
        small = self.mech_eye.main(lua_name="small.lua")
        small = self.rotate_point_cloud(small, 91, "z")
        small = small[small[:,2] > np.min(small[:,2])+37]
        small = small[small[:,0] < np.min(small[:,0])+50]
        small = self.to_origin(small)
        self.rotated_pcd.points = o3d.utility.Vector3dVector(small)
        o3d.io.write_point_cloud("/home/eypan/Documents/alt_jaguar/jaguar_measure/small.ply", self.rotated_pcd)
        
        z_max = np.max(small[:, 2])
        small_z_max_plane = small[small[:, 2] >= z_max - 20]
        small_z_max = np.max(small_z_max_plane[:, 1])
        
        circle_fitter = CircleFitter(small)
        _, z_center_small, radius_small = circle_fitter.fit_circles_and_plot(find_second_circle=False, val_x=0.18, val_z=0.2, delta_z=25)
        s_datum = circle_fitter.get_datum()
        self.dist_3mm_s, _ = circle_fitter.get_distance(second_crc=False, z_distance_to_datum=23.1)
        self.feature_3 = (z_center_small - s_datum)
        self.radius_small = radius_small
        self.z_center_small = z_center_small
        
        return True

    def process_horizontal(self):
        horizontal = self.mech_eye.main(lua_name="horizontal.lua")
        
        horizontal2 = self.mech_eye.main(lua_name="horizontal2.lua")
        
        horizontal2[:,2] -= 70
        horizontal2[:,0] += 50
        self.rotated_pcd.points = o3d.utility.Vector3dVector(horizontal2)
        o3d.io.write_point_cloud("/home/eypan/Documents/alt_jaguar/jaguar_measure/horizontal2.ply", self.rotated_pcd)
        
        # Yatay veri işleme
        diff = np.abs(np.max(horizontal2[:,0])-np.min(horizontal[:,0]))
        datum_horizontal = np.min(horizontal[:,0])+diff
        
        # Çizgi oluşturma ve veri işleme
        y_values = np.linspace(np.min(horizontal[:, 1]), np.max(horizontal[:, 1]), num=100)
        line_points = np.array([[datum_horizontal, y, np.min(horizontal[:, 2])] for y in y_values])
        augmented_point_cloud = np.vstack((horizontal, line_points))
        
        horizontal = self.rotate_point_cloud(augmented_point_cloud, -90, "z")
        horizontal = self.to_origin(horizontal)
        self.horizontal = horizontal
        self.rotated_pcd.points = o3d.utility.Vector3dVector(horizontal)
        o3d.io.write_point_cloud("/home/eypan/Documents/alt_jaguar/jaguar_measure/horizontal_post.ply", self.rotated_pcd)
        
        self.circle_fitter = CircleFitter(horizontal)
        _, circle2 = self.circle_fitter.fit_circles_and_plot()
        self.dist_3mm_h = self.circle_fitter.get_distance()[0]
        self.l_40 = get_40(horizontal)
        self.height = np.max(self.horizontal[:,1])-self.circle_fitter.get_datum()
        self.feature_1 = circle2[1]
        self.feature_2 = circle2[2]
        return True

    def process_vertical(self):
        vertical = self.mech_eye.main(lua_name="vertical.lua")
        vertical = self.remove_gripper_points(vertical)
        
        self.rotated_pcd.points = o3d.utility.Vector3dVector(vertical)
        o3d.io.write_point_cloud("/home/eypan/Documents/alt_jaguar/jaguar_measure/vertical.ply", self.rotated_pcd)
        
        # # Dikey veri işleme
        # x_median = np.median(vertical[:, 0])
        # y_min = np.min(vertical[:, 1])
        # mask = (vertical[:, 1] >= y_min) & (vertical[:, 1] <= y_min + 20)
        # z_max_in_window = np.max(vertical[mask, 2])
        
        # Ölçüm ve hesaplamalar
        l_17_2, ok_17_2 = horn_diff(vertical)
        print("type:",type(ok_17_2))
        l_23_4, ok_23_4 = horn_diff(vertical, 240, 280)
        
        # Robot hareket işlemleri
        current_index = self.get_next_valid_index(self.read_current_point_index(), self.ignored_points, len(self.points))
        self.write_current_point_index(current_index)
        
        if self.same_object and self.cycle < self.range_:
            point = self.points[current_index]
            robot.MoveCart(point["p_up"], 0, 0, vel=100)
            robot.MoveL(point["p"], 0, 0, vel=100)
            robot.WaitMs(500)
            robot.SetDO(7, 0)
            robot.WaitMs(500)
            robot.MoveL(point["p_up"], 0, 0, vel=100)
        else:
            print("DEVAM")
            robot.SetDO(7, 0)
        
        # Son iterasyonda indexi sıfırla
        if self.cycle == self.range_ - 1:  # Son iterasyon
            self.write_current_point_index(0)
        
        # Diğer hesaplamalar
        B = self.circle_fitter.get_B()
        b_trans_val = np.max(vertical[:, 1]) - np.max(self.horizontal[:, 2])
        b_vertical = B + b_trans_val
        
        _, _, r1, l_79_73 = kaydırak(vertical, b_vertical)
        _, _, r2, _ = kaydırak(vertical, y_divisor=0.11, crc_l=27)
        self.r2 = r2
        l_42 = np.max(vertical[:, 1]) - b_vertical
        l_248 = arm_horn_lengths(vertical, b_vertical)
        mean_3mm = np.mean([self.dist_3mm_h, self.dist_3mm_s])
        l_88_6 = (self.feature_1 - self.feature_2)
        l_81_5 = filter_and_visualize_projection_with_ply(self.horizontal)
        
        return {
            "l_17_2": l_17_2,
            "l_23_4": l_23_4,
            "l_42": l_42,
            "l_79_73": l_79_73,
            "l_248": l_248,
            "r1": (r1-self.feature_2),
            "r2": (r2+self.feature_2),
            "feature_1": self.feature_1,
            "mean_3mm": mean_3mm,
            "l_88_6": l_88_6,
            "l_81_5": l_81_5,
            "ok_17_2": ok_17_2
        }

    def run_scan_cycle(self):
        ROBOT_POSITIONS = {
            'scrc': [-500.04, -149.98, 550.01, 82.80, 89.93, -7.30],
            'h_2': [-585.03, 80.01, 444.96, -90, -90, 180],
            'p_91': [-424.99, 49.99, 573.01, -90.00, -0.0005, 90.00],
            'notOK': [-621, 325, 511, 117, -83, -148]
        }
        MAX_RETRIES = 2  # Sabit tanımlandı

        for self.cycle in range(self.range_):
            start_time = time.time()
            current_results = {}
            current_index = self.read_current_point_index()  # Sınıf değişkenine atanabilir
            
            try:
                send_command({"cmd": 107, "data": {"content": "ResetAllError()"}})
                point = self.points[current_index]
                print(f"Cycle {self.cycle+1}/{self.range_}, Processing point: {current_index}")

                self.pick_object(point)
                robot.MoveCart(ROBOT_POSITIONS['scrc'], 0, 0, vel=100)
                
                # current_index ve max_retries parametreleri düzgün iletilmeli
                small_data = self.retry_process(
                    self.process_small, 
                    ROBOT_POSITIONS['scrc'],
                    current_index,  # İlgili parametreye eklendi
                    ROBOT_POSITIONS['notOK'],
                    MAX_RETRIES,
                )
                horizontal_data = self.retry_process(
                    self.process_horizontal, 
                    ROBOT_POSITIONS['h_2'],
                    current_index,
                    ROBOT_POSITIONS['notOK'],
                    MAX_RETRIES,
                )
                vertical_data = self.retry_process(
                    self.process_vertical, 
                    ROBOT_POSITIONS['p_91'],
                    current_index, 
                    ROBOT_POSITIONS['notOK'],
                    MAX_RETRIES,
                )

                current_results = self.combine_results(vertical_data)
                current_results["Processing Time (s)"] = time.time() - start_time

            except Exception as e:
                print(f"Cycle {self.cycle+1} failed with error: {str(e)}")
                current_results["Error"] = str(e)
            
            finally:
                self.results.append(current_results)
        
        self.save_to_excel()

    def retry_process(self, process_func, recovery_position, current_index, trash_position, max_retries=1):  # current_index parametre olarak eklendi
        for attempt in range(max_retries):
            try:
                return process_func()  # process_func'ın dict döndüğünden emin olun
            except Exception as e:
                if attempt == max_retries - 1:
                    updated_index = self.get_next_valid_index(current_index, self.ignored_points, len(self.points))
                    print("updated_index:",updated_index)
                    self.write_current_point_index(updated_index)
                    robot.MoveCart(trash_position, 0, 0, vel=100)
                    robot.SetDO(7, 0)
                    raise

                print(f"Error: {str(e)}\n Retrying {process_func.__name__} after positioning...")
                robot.MoveCart(recovery_position, 0, 0, vel=100)
        
    def combine_results(self, vertical):
        return {
            "Feature1 (102.1)": vertical.get("feature_1"),
            "Feature2 (25mm/2)": self.feature_2,
            "Feature3 (23.1)": self.feature_3,
            "Feature4 (25mm/2)": self.radius_small,
            "Feature5 (L40)": self.l_40,
            "Feature6 (L248)": vertical.get("l_248"),
            "Feature7 (L42)": vertical.get("l_42"),
            "Feature8 (L79.73)": vertical.get("l_79_73"),
            "Feature9 (R1-50)": vertical.get("r1"),
            "Feature10 (R2-35)": vertical.get("r2"),
            "Feature11 (3mm)": vertical.get("mean_3mm"),
            "Feature12 (88.6)": vertical.get("l_88_6"),
            "Feature13 (10.6)": self.feature_3 - self.radius_small,
            "Feature14 (81.5)": vertical.get("l_81_5"),
            "Feature15 (L23.4)": vertical.get("l_23_4"),
            "Feature16 (L17.2)": vertical.get("l_17_2"),
            "Feature17 (2C)": vertical.get("ok_17_2"),
        }


    def save_to_excel(self):
        rows = []
        for iteration, result in enumerate(self.results, start=1):
            for name, value in result.items():
                rows.append({"Iteration": iteration, "Feature": name, "Value": value})
        
        results_df = pd.DataFrame(rows)
        output_file = "/home/eypan/Documents/alt_jaguar/jaguar_measure/scan_results.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, index=False, header=True, sheet_name="ScanResults")
                workbook = writer.book
                worksheet = writer.sheets["ScanResults"]
                
                for row_index, row in enumerate(results_df.itertuples(index=False), start=2):
                    iteration = row.Iteration
                    feature = row.Feature
                    value = row.Value
                    
                    iteration_color = "FCE4D6" if iteration % 2 == 0 else "D9EAD3"
                    iteration_fill = PatternFill(start_color=iteration_color, end_color=iteration_color, fill_type="solid")
                    for col_index in range(1, len(results_df.columns) + 1):
                        worksheet.cell(row=row_index, column=col_index).fill = iteration_fill
                    
                    if feature in self.tolerances:
                        target, tolerance = self.tolerances[feature]
                        col_index = 3  # 'Value' column
                        cell = worksheet.cell(row=row_index, column=col_index)
                        if target - tolerance <= value <= target + tolerance:
                            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        except FileNotFoundError:
            print(f"Error: Could not find the directory for {output_file}. Check the path and try again.")
        
        print("Results successfully saved to Excel with color-coded iterations and tolerance.")

    @staticmethod
    def to_origin(points):
        min_old = np.min(points[:, 1])
        min_old_x = np.min(points[:, 0])
        points[:, 1] += -min_old
        points[:, 0] += -min_old_x
        return points

    def read_current_point_index(self):
        if os.path.exists(self.file_path):
            with open(self.file_path, 'r') as file:
                return int(file.read().strip())
        return 0

    def write_current_point_index(self, index):
        with open(self.file_path, 'w') as file:
            file.write(str(index))


if __name__ == "__main__":
    scanner = JaguarScanner()
    scanner.run_scan_cycle()