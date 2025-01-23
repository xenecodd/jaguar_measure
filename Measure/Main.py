import os
from tkinter.messagebox import showerror
from MecheyePackage.robot_control import send_command
from Scripts import *
from MecheyePackage import TriggerWithExternalDeviceAndFixedRate, robot
import open3d as o3d
import numpy as np
import matplotlib
matplotlib.use('Agg')
import time
import pandas as pd
from openpyxl.styles import PatternFill


if 0 :        
    p1_up = [-747.6768798828125, -36.60947036743164, 224.6449432373047, 179.9124450683593, 0.01771553605794907, 90.00694274902342]
    p1    = [-747.6768188476562, -36.60196685791015, 139.29833984375, 179.9137420654297, 0.016725072637200356, 88.00242614746092]   
    p2_up = [-629, -36.60947036743164, 224.6449432373047, 179.9124450683593, 0.01771553605794907, 88.00694274902342]
    p2    = [-629, -36.60947036743164, 139.29833984375,  179.9124450683593, 0.01771553605794907, 88.00694274902342]
    p3_up = [-519, -58.2884407043457, 220.5605010986328, 179.9501495361328, 0.01032531540840864, 88.00717163085938]
    p3    = [-519, -58.29881286621093, 135.5182800292968, 179.9518737792968, 0.010120976716279985, 88.0091323852539]
    p4_up = [-398.1421813964843, -58.29881286621093, 220.5605010986328, 179.9518737792968, 0.010120976716279985, 88.0091323852539]
    p4   = [-398.1421813964843, -58.2884407043457, 135.5182800292968, 179.9501495361328, 0.01032531540840864, 88.00717163085938]

def rotate_point_cloud(points, angle_degrees, rot):
    """
    Bir nokta bulutunu Z ekseni etrafında belirli bir açıyla döndür.
    :param points: Nx3 boyutunda numpy dizisi (nokta bulutu)
    :param angle_degrees: Döndürme açısı (derece cinsinden)
    :return: Döndürülmüş nokta bulutu (Nx3 numpy dizisi)
    """
    # Dereceden radyana çevir
    angle_radians = np.radians(angle_degrees)
    
    # Z ekseni etrafında dönüşüm matrisi
    if rot == "z":
        
        rotation_matrix = np.array([
        [np.cos(angle_radians), -np.sin(angle_radians), 0],
        [np.sin(angle_radians),  np.cos(angle_radians), 0],
        [0, 0, 1]
    ])
    elif rot == "x":

        rotation_matrix = np.array([
        [1, 0, 0],
        [0, np.cos(angle_radians), -np.sin(angle_radians)],
        [0, np.sin(angle_radians), np.cos(angle_radians)]
    ])
    elif rot == "y":
            
        rotation_matrix = np.array([
        [np.cos(angle_radians), 0, np.sin(angle_radians)],
        [0, 1, 0],
        [-np.sin(angle_radians), 0, np.cos(angle_radians)]
    ])
    
    # Noktaları döndür
    rotated_points = points @ rotation_matrix.T
    return rotated_points

def save_point_cloud(pcd, file_path):
    # Ensure that pcd is a PointCloud object, not a numpy array
    if isinstance(pcd, o3d.geometry.PointCloud):
        pcd_array = np.asarray(pcd.points)  # Convert PointCloud points to numpy array
    else:
        pcd_array = pcd  # If pcd is already a numpy array, use it directly

    # Create a new PointCloud and set the points
    pcd_new = o3d.geometry.PointCloud()
    pcd_new.points = o3d.utility.Vector3dVector(pcd_array)

    # Write the point cloud to the file
    o3d.io.write_point_cloud(file_path, pcd_new)

def remove_gripper_points(points):
    """
    Belirtilen Y ekseni aralığında yer alan noktaları bir nokta bulutundan siler.

    Args:
        point_cloud (o3d.geometry.PointCloud): Giriş nokta bulutu.
        y_min (float): Silinecek aralığın alt sınırı (dahil).
        y_max (float): Silinecek aralığın üst sınırı (dahil).

    Returns:
        o3d.geometry.PointCloud: Filtrelenmiş nokta bulutu.
    """
    min_x = np.min(points[:, 0])
    
    y_min, y_max = np.min(points[:,1][points[:,0] < min_x+23]), np.max(points[:,1][points[:,0] < min_x+23])
    # Y ekseni değerine göre filtreleme
    filtered_points = points[(points[:, 1] < y_min) | (points[:, 1] > y_max)]

    return filtered_points

def to_origin(points):
    """Nokta bulutunu orijine tasi"""
    min_old = np.min(points[:, 1])
    min_old_x = np.min(points[:, 0])
    points[:, 1] += -min_old
    points[:, 0] += -min_old_x
    return points

def clean_point_cloud_numpy(points, nb_neighbors=20, std_ratio=2.0):
    """
    Gürültülü noktaları temizlemek için bir numpy nokta bulutunu filtreler.

    Parameters:
        points (numpy.ndarray): (N, 3) boyutunda nokta bulutu verisi.
        nb_neighbors (int): Her bir nokta için kullanılacak komşu sayısı.
        std_ratio (float): Gürültü temizleme için standart sapma oranı.

    Returns:
        numpy.ndarray: Temizlenmiş nokta bulutu.
    """
    from sklearn.neighbors import NearestNeighbors

    # Komşuluk analizi için k-d ağacı
    nbrs = NearestNeighbors(n_neighbors=nb_neighbors).fit(points)
    distances, _ = nbrs.kneighbors(points)

    # Komşuluk mesafelerinin ortalama değerini hesapla
    mean_distances = np.mean(distances[:, 1:], axis=1)  # İlk sütun kendisi olduğu için hariç tutulur

    # Ortalama mesafelerin global ortalaması ve standart sapması
    mean_global = np.mean(mean_distances)
    std_global = np.std(mean_distances)

    # Gürültü dışındaki noktaları belirle
    mask = mean_distances <= mean_global + std_ratio * std_global

    # Temizlenmiş nokta bulutunu döndür
    cleaned_points = points[mask]
    print(f"{len(points) - len(cleaned_points)} gürültü noktası kaldırıldı.")
    return cleaned_points

# Initialize Mech-Eye profiler with external device and fixed rate trigger
mech_eye = TriggerWithExternalDeviceAndFixedRate()

rotated_pcd = o3d.geometry.PointCloud()

# Data storage list for Excel
results = []

# Noktalar
points = [
    {"p_up": [-747.6768798828125, -36.60947036743164, 224.6449432373047, 179.9124450683593, 0.01771553605794907, 90.00694274902342], 
     "p": [-747.6768188476562, -36.60196685791015, 139.29833984375, 179.9137420654297, 0.016725072637200356, 88.00242614746092]},
    {"p_up": [-629, -36.60947036743164, 224.6449432373047, 179.9124450683593, 0.01771553605794907, 88.00694274902342], 
     "p": [-629, -36.60947036743164, 139.29833984375, 179.9124450683593, 0.01771553605794907, 88.5]},
    {"p_up": [-519, -58.2884407043457, 220.5605010986328, 179.9501495361328, 0.01032531540840864, 88.00717163085938], 
     "p": [-519, -58.29881286621093, 135.5182800292968, 179.9518737792968, 0.010120976716279985, 88.0091323852539]},
    {"p_up": [-398.1421813964843, -58.29881286621093, 220.5605010986328, 179.9518737792968, 0.010120976716279985, 88.0091323852539], 
     "p": [-398.1421813964843, -58.2884407043457, 135.5182800292968, 179.9501495361328, 0.01032531540840864, 88.00717163085938]}
]

# Nokta Indeksi Dosyası
file_path = "/home/eypan/Documents/scanner/point_index.txt"

# Dosyadan mevcut noktayı oku
def read_current_point_index():
    if os.path.exists(file_path):
        with open(file_path, 'r') as file:
            return int(file.read().strip())
    return 0  # Başlangıçta sıfırdan başla

# Gelecek noktayı kaydet
def write_current_point_index(index):
    with open(file_path, 'w') as file:
        file.write(str(index))
j = read_current_point_index()
range_=7
same_object =False
for i in range(range_):
    start_time = time.time()
    send_command({"cmd": 107, "data": {"content": "ResetAllError()"}})
    if i == 0:
        i=i+j
    point = points[i%4]

    # İlk hareket
    ret = robot.MoveCart(point["p_up"], 0, 0, vel=100)
    ret = robot.MoveL(point["p"], 0, 0, vel=100)
    robot.WaitMs(500)
    robot.SetDO(7, 1)
    robot.WaitMs(500)
    ret = robot.MoveL(point["p_up"], 0, 0, vel=100)

    scrc = [-500.0421752929687, -149.981979370117, 550.0119018554688, 82.80037689208984, 89.9276657104492, -7.29904794692993]
    ret = robot.MoveCart(scrc, 0, 0, vel=100)
    small = mech_eye.main(lua_name="small.lua")
    # small=clean_point_cloud_numpy(small)
    
    small = rotate_point_cloud(small, 90, "z")
    # small = rotate_point_cloud(small, 2, "x")
    small= small[small[:,2]>np.min(small[:,2])+37]
    small= small[small[:,0]<np.min(small[:,0])+50] # gripper filterleniyor bu kısımda
    small= small[small[:,0]>np.min(small[:,0])+1.5]
    small = to_origin(small)
    rotated_pcd.points = o3d.utility.Vector3dVector(small)
    o3d.io.write_point_cloud("/home/eypan/Documents/scanner/jaguar_measure/small.ply", rotated_pcd)


    z_max = np.max(small[:, 2])
    small_z_max_plane = small[small[:, 2] >= z_max - 20]
    small_z_max = np.max(small_z_max_plane[:, 1])   # small_z_max tanımlandığı konum önemli
    # plt.plot(small_z_max_plane[:, 0], small_z_max_plane[:, 1], 'bo')

    circle_fitter = CircleFitter(small) 



    _,z_center_small,radius_small = circle_fitter.fit_circles_and_plot(find_second_circle=False,val_x=0.18,val_z=0.2,delta_z=25)
    s_datum = circle_fitter.get_datum() # s_datum değişkeninin tanımlandığı yerin konumu önemli
    dist_3mm_s,_ = circle_fitter.get_distance(second_crc=False, z_distance_to_datum=23.1)
    print("s_datum", s_datum)
    feature_3 = (z_center_small- s_datum)

    horizontal = mech_eye.main(lua_name="horizontal.lua")

    #///////////////////////////////////////////////////////////////////
    horizontal2 = mech_eye.main(lua_name="horizontal2.lua")
    
    horizontal2[:,2] -= 70
    horizontal2[:,0] += 50
    rotated_pcd.points = o3d.utility.Vector3dVector(horizontal2)
    o3d.io.write_point_cloud("/home/eypan/Documents/scanner/jaguar_measure/horizontal2.ply", rotated_pcd)
    print("KAYDETTİK HORIZONTAL2")
    
    #///////////////////////////////////////////////////////////////////





    diff = np.abs(np.max(horizontal2[:,0])-np.min(horizontal[:,0]))
    print("diff",diff)
    datum_horizontal = np.min(horizontal[:,0])+diff
    print("datum_horizontal",datum_horizontal)

    # Çizgi için parametreler
    x_coord = datum_horizontal  # Çizginin sabit x ekseni koordinatı
    y_start = np.min(horizontal[:, 1])  # Y ekseni boyunca başlangıç noktası
    y_end = np.max(horizontal[:, 1])  # Y ekseni boyunca bitiş noktası
    z_coord = np.min(horizontal[:, 2])  # Çizginin sabit z ekseni koordinatı

    # Çizgi noktalarını oluştur
    y_values = np.linspace(y_start, y_end, num=100)  # Y ekseni boyunca 100 ara nokta
    line_points = np.array([[x_coord, y, z_coord] for y in y_values])

    # Çizgiyi nokta bulutuna ekle
    augmented_point_cloud = np.vstack((horizontal, line_points))

    rotated_pcd.points = o3d.utility.Vector3dVector(horizontal)
    o3d.io.write_point_cloud("/home/eypan/Documents/alt_jaguar/jaguar_measure/horizontal_pre.ply", rotated_pcd)
    horizontal = rotate_point_cloud(augmented_point_cloud, -90, "z")
    horizontal = to_origin(horizontal)
    rotated_pcd.points = o3d.utility.Vector3dVector(horizontal)
    o3d.io.write_point_cloud("/home/eypan/Documents/alt_jaguar/jaguar_measure/horizontal_post.ply", rotated_pcd)
    l_40 = get_40(horizontal)

    circle_fitter = CircleFitter(horizontal)
    circle, circle2= circle_fitter.fit_circles_and_plot()
    
    feature_2 = circle2[2]
    feature_2_z_coord = circle2[1]
    vertical = mech_eye.main(lua_name="vertical.lua")
    # vertical=clean_point_cloud_numpy(vertical)
    if same_object and i<range_:
        point = points[i+1]
        ret = robot.MoveCart(point["p_up"], 0, 0, vel=100)
        ret = robot.MoveL(point["p"], 0, 0, vel=100)
        robot.WaitMs(500)
        robot.SetDO(7, 0)
        robot.WaitMs(500)
        ret = robot.MoveL(point["p_up"], 0, 0, vel=100)
    else :
        robot.SetDO(7, 0)
    write_current_point_index(i + 1)
    if i ==range_-1:
        write_current_point_index(0)
    vertical = remove_gripper_points(vertical)
    rotated_pcd.points = o3d.utility.Vector3dVector(vertical)
    o3d.io.write_point_cloud("/home/eypan/Documents/scanner/jaguar_measure/vertical.ply", rotated_pcd)
    vertical_copy = vertical.copy()

    x_median = np.median(vertical[:, 0])
    y_min = np.min(vertical[:, 1])

    mask = (
        (vertical[:, 1] >= y_min) & (vertical[:, 1] <= y_min + 20)
    )
    z_max_in_window = np.max(vertical[mask, 2])

    
    datum_trans_val =(z_max_in_window - small_z_max )
    datum_vertical = s_datum + datum_trans_val
    print("datum vertical",datum_vertical)
    l_17_2, ok_17_2 = horn_diff(vertical)
    l_23_4, ok_23_4 = horn_diff(vertical, 240, 280)

    B = circle_fitter.get_B()
    b_trans_val = np.max(vertical[:, 1]) - np.max(horizontal[:, 2])
    b_vertical = B + b_trans_val


    _, center_z, radius_big = bigcircle(vertical, crc=True)
    _, _, small_rad = bigcircle(vertical, "small_circle", val_x=0.90, val_y=0.2, val_z=0.05, crc=True)
    print("small_rad",small_rad)
    _, _, r1, l_79_73 = kaydırak(vertical, b_vertical)

    # z_trans_val = circle[1] - center_z
    # datum = circle_fitter.get_datum()
    # datum_vertical = datum - z_trans_val

    _, _, r2, _ = kaydırak(vertical, y_divisor=0.11, crc_l=27)

    l_42 = np.max(vertical[:, 1]) - b_vertical
    l_248 = arm_horn_lengths(vertical_copy, b_vertical)
    height=np.max(horizontal[:,1])-circle_fitter.get_datum()
    print("height", height)
    print("datum_horizontal", datum_horizontal)
    dist_3mm_h, _ = circle_fitter.get_distance()
    feature_1 = np.abs(circle_fitter.get_datum()-feature_2_z_coord)
    mean_3mm = np.mean([dist_3mm_h, dist_3mm_s])
    l_88_6,l_81_5 = filter_and_visualize_projection_with_ply(horizontal,circle_fitter.get_datum())
    processing_time = time.time() - start_time
    # Append results for this iteration

    # Append new row to results
    results.append({
        "Feature1 (102.1)": feature_1,
        "Feature2 (25mm/2)": feature_2,
        "Feature3 (23.1)": feature_3,
        "Feature4 (25mm/2)": radius_small,
        "Feature5 (L40)": l_40,
        "Feature6 (L248)": l_248,
        "Feature7 (L42)": l_42,
        "Feature8 (L79.73)": l_79_73,
        "Feature9 (R1-62.5)": r1,
        "Feature10 (R2-22.5)": r2,
        "Feature11 (3mm)": mean_3mm,
        "Feature12 (88.6)": l_88_6,
        "Feature13 (10.6)": (feature_3 - radius_small),
        "Feature14 (81.5)": l_81_5,
        "Feature15 (L23.4)": l_23_4,
        "Feature16 (L17.2)": l_17_2,
        "Feature17 (2C)": ok_17_2,
        "Processing Time (s)": processing_time
    })

# Define tolerance values
tolerances = {
    "Feature1 (102.1)": (102.1, 1.5),
    "Feature2 (25mm/2)": (12.5, 0.5),
    "Feature3 (23.1)": (23.1, 1),
    "Feature4 (25mm/2)": (12.5, 0.5),
    "Feature5 (L40)": (40.0, 1.5),
    "Feature6 (L248)": (248.0, 2.0),
    "Feature7 (L42)": (42.0, 1.5),
    "Feature8 (L79.73)": (79.73, 1.5),
    "Feature9 (R1-62.5)": (62.5, 1.5),
    "Feature10 (R2-22.5)": (22.5, 1),
    "Feature11 (3mm)": (0, 1.5),
    "Feature12 (88.6)": (88.6, 1.5),
    "Feature13 (10.6)": (10.6, 1.0),
    "Feature14 (81.5)": (81.5, 1.5),
    "Feature15 (L23.4)": (23.4, 1.0),
    "Feature16 (L17.2)": (17.2, 1.0),
}

# Transform results for Excel
rows = []
for iteration, result in enumerate(results, start=1):
    for name, value in result.items():
        rows.append({"Iteration": iteration, "Feature": name, "Value": value})

results_df = pd.DataFrame(rows)

# Output file path
output_file = "/home/eypan/Documents/scanner/jaguar_measure/scan_results.xlsx"
os.makedirs(os.path.dirname(output_file), exist_ok=True)  # Ensure directory exists

try:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False, header=True, sheet_name="ScanResults")

        # Access workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets["ScanResults"]

        # Apply color coding
        for row_index, row in enumerate(results_df.itertuples(index=False), start=2):
            iteration = row.Iteration
            feature = row.Feature
            value = row.Value

            # Apply iteration-based color (alternating colors for rows)
            iteration_color = "FCE4D6" if iteration % 2 == 0 else "D9EAD3"  # Light red or green
            iteration_fill = PatternFill(start_color=iteration_color, end_color=iteration_color, fill_type="solid")
            for col_index in range(1, len(results_df.columns) + 1):
                worksheet.cell(row=row_index, column=col_index).fill = iteration_fill

            # Apply tolerance-based color
            if feature in tolerances:
                target, tolerance = tolerances[feature]
                col_index = 3  # 'Value' column
                cell = worksheet.cell(row=row_index, column=col_index)
                if target - tolerance <= value <= target + tolerance:
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

except FileNotFoundError:
    print(f"Error: Could not find the directory for {output_file}. Check the path and try again.")

print("Results successfully saved to Excel with color-coded iterations and tolerance.")
