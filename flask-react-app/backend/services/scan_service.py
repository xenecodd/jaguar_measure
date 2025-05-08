import subprocess
import threading
import time
import json
import numpy as np
import logging
from backend.models.robot_state import state
from backend.config import Config
from services.robot_service import safe_get_di
import multiprocessing
from backend.config import Config
from Measure.MecheyePackage.mecheye_trigger import robot
from openpyxl.styles import PatternFill
import os
import pandas as pd
from pathlib import Path

logger = logging.getLogger(__name__)

with open(Config.CONFIG_PATH, "r") as f:
    config = json.load(f)

def calculate_tolerance_distance(value: float, target: float) -> float:
    """Calculate the absolute distance between a value and its target."""
    return np.abs(value - target)

def get_gradient_color(distance: float, tolerance: float) -> str:
    """
    Generate a color hex code based on how far a value is from its tolerance range.
    
    Args:
        distance: The absolute distance from the target value
        tolerance: The allowed tolerance range
    
    Returns:
        A hex color code string (e.g., "FF0000" for red)
    """
    if distance > tolerance:
        return "FF0000"  # Red for values outside tolerance
    
    ratio = distance / tolerance
    red = int(255 * ratio)
    green = int(255 * (1 - ratio))
    blue = 0
    
    return f"{red:02X}{green:02X}{blue:02X}"

def make_json_serializable(obj):
    if isinstance(obj, dict):
        return {k: make_json_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [make_json_serializable(v) for v in obj]
    elif isinstance(obj, (np.bool_,)):  # numpy bool varsa
        return bool(obj)
    elif isinstance(obj, (np.integer, np.int64, np.float64)):
        return obj.item()  # numpy sayıları
    else:
        return obj

def generate_json_data(results, tolerances=config['tolerances']):
    """
    Generates JSON data from scan results, with tolerance checks and color coding.
    
    Args:
        results (list): List of dictionaries containing scan result data
        tolerances (dict, optional): Dictionary of feature tolerances with format {feature: (target, tolerance)}
    
    Returns:
        dict: JSON-compatible data structure containing processed results with color coding
    """
    if tolerances is None:
        tolerances = {}
    
    processed_data = {
        "scan_results": [],
        "summary": {
            "total_iterations": len(results),
            "passed_iterations": 0
        }
    }
    
    for iteration, result in enumerate(results, start=1):
        iteration_ok = True
        
        # Determine iteration background color (alternating)
        iteration_color = "FCE4D6" if iteration % 2 == 0 else "D9EAD3"
        
        iteration_data = {
            "iteration": iteration,
            "background_color": iteration_color,
            "features": [],
            "status": {
                "ok": True,
                "color": "00B050"  # Default to green (will be updated if any test fails)
            }
        }
        
        for feature, value in result.items():
            feature_data = {
                "name": feature,
                "value": value,
                "background_color": iteration_color
            }
            
            # Handle dictionary values
            if isinstance(value, dict):
                feature_data["value"] = value
            
            # Handle non-serializable values
            try:
                json.dumps(feature_data["value"])
            except TypeError:
                feature_data["value"] = str(value)
            
            # Tolerance check and color coding
            if feature in tolerances:
                target, tolerance = tolerances[feature]
                
                try:
                    # Try to convert to numeric for tolerance check
                    numeric_value = float(value) if isinstance(value, str) else value
                    
                    # Calculate distance from target
                    distance = calculate_tolerance_distance(numeric_value, target)
                    within_tolerance = distance <= tolerance
                    
                    # Generate color code
                    color_code = get_gradient_color(distance, tolerance)
                    
                    tolerance_data = {
                        "target": target,
                        "tolerance": tolerance,
                        "distance": distance,
                        "tolerance_remaining": tolerance - distance,
                        "within_tolerance": within_tolerance,
                        "color": "00B050" if within_tolerance else "FF0000"  # Green if within tolerance, red if not
                    }
                    
                    # Add gradient color for visualization
                    tolerance_data["gradient_color"] = color_code
                    
                    feature_data["tolerance_check"] = tolerance_data
                    feature_data["value_color"] = tolerance_data["color"]
                    
                    # Update iteration status if tolerance check fails
                    if not within_tolerance:
                        iteration_data["status"]["ok"] = False
                        iteration_data["status"]["color"] = "FF0000"  # Red for failed iteration
                        if "failure_reasons" not in iteration_data["status"]:
                            iteration_data["status"]["failure_reasons"] = []
                        iteration_data["status"]["failure_reasons"].append(f"{feature} out of tolerance")
                        
                except (ValueError, TypeError):
                    # If value can't be converted to numeric, no tolerance check possible
                    feature_data["tolerance_check"] = {
                        "error": "Value cannot be numerically compared",
                        "target": target,
                        "tolerance": tolerance
                    }
            
            iteration_data["features"].append(feature_data)
        
        # Update summary data
        if iteration_data["status"]["ok"]:
            processed_data["summary"]["passed_iterations"] += 1
            
        processed_data["scan_results"].append(iteration_data)
    
    # Add pass rate to summary
    processed_data["summary"]["pass_rate"] = (
        processed_data["summary"]["passed_iterations"] / processed_data["summary"]["total_iterations"] 
        if processed_data["summary"]["total_iterations"] > 0 else 0
    )
    
    # Add color coding to summary
    pass_rate = processed_data["summary"]["pass_rate"]
    if pass_rate == 1.0:
        processed_data["summary"]["color"] = "00B050"  # Green for 100% pass
    elif pass_rate >= 0.8:
        processed_data["summary"]["color"] = "FFFF00"  # Yellow for 80%+ pass
    else:
        processed_data["summary"]["color"] = "FF0000"  # Red for <80% pass
    
    # Make JSON serializable
    serializable_data = make_json_serializable(processed_data)
    
    # Save to file for debugging
    try:
        with open("processed_data.json", "w", encoding="utf-8") as f:
            json.dump(serializable_data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Error saving processed data to file: {e}")
    
    return serializable_data

def save_to_excel(results, tolerances=config['tolerances']):
    rows = []
    for iteration, result in enumerate(results, start=1):
        for name, value in result.items():
            rows.append({"Iteration": iteration, "Feature": name, "Value": value})

    results_df = pd.DataFrame(rows)

    output_file = Path(__file__).resolve().parent.parent / 'excel' / 'scan_results.xlsx'
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, header=True, sheet_name="ScanResults")
            worksheet = writer.sheets["ScanResults"]

            for row_index, row in enumerate(results_df.itertuples(index=False), start=2):
                iteration = row.Iteration
                feature = row.Feature
                value = row.Value

                # 1) Iteration bazında satır boyama
                iteration_color = "FCE4D6" if iteration % 2 == 0 else "D9EAD3"
                iteration_fill = PatternFill(start_color=iteration_color, end_color=iteration_color, fill_type="solid")
                for col_index in range(1, len(results_df.columns) + 1):
                    worksheet.cell(row=row_index, column=col_index).fill = iteration_fill

                # 2) Tolerans kontrolü: Value hücresini (3. sütun) yeşil/kırmızı boyama
                if feature in tolerances:
                    target, tolerance = tolerances[feature]
                    col_index = 3  # 'Value' sütunu
                    value_cell = worksheet.cell(row=row_index, column=col_index)

                    if target - tolerance <= value <= target + tolerance:
                        value_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    else:
                        value_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    # 3) Distance hücresine yaz ve renklendir
                    distance = calculate_tolerance_distance(value, target)
                    gradient_color = get_gradient_color(distance, tolerance)

                    distance_cell = worksheet.cell(row=row_index, column=col_index + 1)
                    distance_cell.value = round(tolerance - distance, 4)
                    distance_cell.fill = PatternFill(start_color=gradient_color,
                                                     end_color=gradient_color,
                                                     fill_type="solid")

        print("Results successfully saved to Excel with color-coded iterations and tolerance.")

    except FileNotFoundError:
        print(f"Error: Could not find the directory for {output_file}. Check the path and try again.")


def log_subprocess_output(process, log_prefix="[SCAN]", error_prefix="[SCAN_ERR]"):
    """
    Log subprocess stdout and stderr in separate threads
    
    Args:
        process: Subprocess.Popen object to monitor
        log_prefix: Prefix for stdout log messages
        error_prefix: Prefix for stderr log messages
    """
    def _log_stdout():
        try:
            for line in iter(process.stdout.readline, ''):
                if line:
                    logger.info(f"{log_prefix} {line.strip()}")
                    pass
                else:
                    break
        except Exception as e:
            logger.error(f"Error logging stdout: {e}")

    def _log_stderr():
        try:
            for line in iter(process.stderr.readline, ''):
                if not line:
                    break
                # Filter out specific error messages
                if "Failed to obtain the IP address" in line:
                    continue
                logger.error(f"{error_prefix} {line.strip()}")
        except Exception as e:
            logger.error(f"Error logging stderr: {e}")
            
    # Start logging threads
    stdout_thread = threading.Thread(target=_log_stdout, daemon=True)
    stderr_thread = threading.Thread(target=_log_stderr, daemon=True)
    stdout_thread.start()
    stderr_thread.start()
    
    return stdout_thread, stderr_thread

def cleanup_scan_process(process):
    """Safely clean up a subprocess"""
    if process:
        try:
            if process.stdin:
                process.stdin.close()
            if process.poll() is None:  # Still running
                process.terminate()
                try:
                    process.wait(timeout=5)  # Wait up to 5 seconds
                except subprocess.TimeoutExpired:
                    process.kill()  # Force kill if terminate hangs
                    process.wait()
            logger.info("scan.py was closed")
        except Exception as e:
            logger.error(f"Error closing scan process: {e}")

def monitor_robot(stop_event, restart_event):
    """
    Robot DI değerlerine göre taramayı durdurup yeniden başlatmak için kontrol.
    stop_event: Scan sürecinin durdurulması için event
    restart_event: Yeniden başlatma sinyali için event
    """
    # Başlatmadan önce kısa bir bekleme süresi
    time.sleep(1)
    
    logger.info(f"Monitor robot started - Current DI8: {safe_get_di(8, 0)}")
    
    # DI8 durumunu kontrol et; güvenli değilse scan durduruluyor.
    while not stop_event.is_set():
        try:
            di8 = safe_get_di(8, 0)
        except Exception as e:
            logger.error(f"Error reading DI8: {e}")
            break

        if di8 != (0, 0):
            logger.info(f"Stop condition met: DI8={di8}")
            state.profiler.stop_acquisition()
            state.profiler.disconnect()
            robot.StopMotion()
            stop_event.set()
            state.set_scan_started(False)
            break
            
        time.sleep(0.5)
    
    logger.info("Monitor robot waiting for restart signal...")
    # DI9 üzerinden yeniden başlatma sinyali bekleniyor:
    wait_time = 5
    start = time.time()
    while True:
        try:
            current_di9 = safe_get_di(9, 0)
        except Exception as e:
            logger.error(f"Error reading DI9: {e}")
            break

        if current_di9 == (0, 1):
            logger.info(f"Restart signal detected: DI9={current_di9}")
            restart_event.set()
            break
        elif time.time() - start > wait_time:
            logger.info("Waiting for restart signal. Please press start.")
            wait_time += 5
            start = time.time()
    
    logger.info("Monitor robot exiting.")

def auto_restart_monitor():
    """
    Monitor for automatic restart conditions based on DI values
    """
    state.set_auto_monitor_running(True)
    logger.info("Auto-restart monitor started")

    while True:
        # Check conditions for starting a new scan:
        # 1. Scan is not already running
        # 2. DI9 (start button) is (0,1)
        # 3. DI8 (safety button) is (0,0)
        if (not state.scan_started and 
            state.di9_status == (0, 1) and 
            state.di8_status == (0, 0)):
            
            logger.info(f"Auto-restart triggered by DI9={state.di9_status}, DI8={state.di8_status}")

            # Create new events
            state.stop_event = multiprocessing.Event()
            state.restart_event = multiprocessing.Event()

            # Start new scan process
            scan_process_thread = threading.Thread(
                target=run_scan, 
                args=(state.stop_event,)
            )
            scan_process_thread.start()

            # Start new monitor thread
            state.monitor_thread = threading.Thread(
                target=monitor_robot, 
                args=(state.stop_event, state.restart_event), 
                daemon=True
            )
            state.monitor_thread.start()

            state.set_scan_started(True)
            logger.info("Scan started successfully")

            # Wait for initialization
            time.sleep(1)

        # Check for restart request
        if state.restart_event and state.restart_event.is_set() and not state.scan_started:
            logger.info("Processing restart event")
            state.restart_event.clear()

            if state.di8_status == (0, 0):
                logger.info("Conditions are safe for restart")
                state.stop_event = multiprocessing.Event()

                scan_process_thread = threading.Thread(
                    target=run_scan, 
                    args=(state.stop_event,)
                )
                scan_process_thread.start()

                state.monitor_thread = threading.Thread(
                    target=monitor_robot, 
                    args=(state.stop_event, state.restart_event), 
                    daemon=True
                )
                state.monitor_thread.start()

                state.set_scan_started(True)
                logger.info("Scan restarted successfully")
            else:
                logger.warning(f"Cannot restart: unsafe conditions DI8={state.di8_status}")

        time.sleep(0.5)

def run_scan(stop_event):
    """
    Function that starts scan.py subprocess with proper input/output handling
    
    Args:
        stop_event: Event to signal when to stop scanning
    """
    try:
        # Start the subprocess with stdin as a pipe
        scan_process = subprocess.Popen(
            ["python3", Config.SCAN_SCRIPT_PATH],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE,
            bufsize=1,
            universal_newlines=True
        )
        logger.info("Scan process started successfully")
        
        # Update global state
        state.set_scan_process(scan_process)
        
        # Start logging threads
        stdout_thread, stderr_thread = log_subprocess_output(scan_process)
        
        # Monitor subprocess to prevent zombies
        start_time = time.time()
        
        while not stop_event.is_set():
            # Check if process has terminated unexpectedly
            if scan_process.poll() is not None:
                logger.warning(f"Scan process terminated with code {scan_process.returncode}")
                break
                
            # Check if process has been running too long (timeout)
            if time.time() - start_time > 3600:  # 1 hour timeout
                logger.warning("Scan process timeout - terminating")
                break
                
            time.sleep(0.1)

        # Stop signal received, terminate subprocess
        cleanup_scan_process(scan_process)
        
        # Wait for logging threads to finish
        stdout_thread.join(timeout=1)
        stderr_thread.join(timeout=1)
        
        return scan_process.stdout, scan_process.stderr
    except Exception as e:
        logger.error(f"Error in run_scan: {e}")
        return None, None
