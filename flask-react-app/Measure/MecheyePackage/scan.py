import os
import threading
import time
import numpy as np
import open3d as o3d
import pandas as pd
import mysql.connector
import matplotlib
from openpyxl.styles import PatternFill
from robot_control import send_command
from Measure.Scripts import *
from mecheye_trigger import TriggerWithExternalDeviceAndFixedRate, robot
from points import *
from config import config
import sys
import json

# Backend yapılandırması: Agg mod, non-interaktif; TkAgg ise interaktif.
if config.use_agg:
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    plt.show = lambda: None
else:
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt

class JaguarScanner:
    """
    JaguarScanner sınıfı; robot kontrolü, tarama ve ölçüm işlemlerini yürüten profesyonel bir uygulamadır.
    
    Özellikler:
      - Ölçüm verilerini toplar, işler ve Excel dosyasına kaydeder.
      - Robotun pick (parça alma) ve put-back (parça geri koyma) işlemlerini yönetir.
      - Thread'li hesaplama ile performansı artırır.
    """
    def __init__(self, vel_mul: float=1, use_agg: bool = config.use_agg, put_back: bool = False):
        """
        JaguarScanner örneğini başlatır.
        
        Args:
            vel_mul (float): Hız çarpanı.
            use_agg (bool, optional): Agg mod kullanımı. Varsayılan True.
            put_back (bool, optional): Parçanın geri konulup konulacağı. False ise trash noktasına yönlendirilir.
        """
        self.config = config
        self.mech_eye = TriggerWithExternalDeviceAndFixedRate(vel_mul)
        self.pcd = o3d.geometry.PointCloud()
        self.results = []
        self.excel_threads = []
        self.points = left_of_robot_points + left_small #+ back_points + right_of_robot_points + right_small
        self.pick_point = 1
        self.current_di0_value = (0, 0)  # Varsayılan değerle başlat
        self.di0_thread = threading.Thread(target=self.read_di0_updates, daemon=True)
        self.di0_thread.start()
        self.tolerances = {
            "Feature1 (102.1)": (102.1, 2.0),
            "Feature2 (25mm/2)": (12.5, 0.5),
            "Feature3 (23.1)": (23.1, 1),
            "Feature4 (25mm/2)": (12.5, 0.5),
            "Feature5 (L40)": (40.0, 1),
            "Feature6 (L248)": (248.0, 2.0),
            "Feature7 (L42)": (42.0, 1.5),
            "Feature8 (L79.73)": (79.73, 1.5),
            "Feature9 (R1-50)": (50, 1.5),
            "Feature10 (R2-35)": (35, 1.5),
            "Feature11 (3mm)": (0, 3.0),
            "Feature12 (88.6)": (88.6, 1.5),
            "Feature13 (10.6)": (10.6, 1.0),
            "Feature14 (81.5)": (81.5, 1.5),
            "Feature15 (L23.4)": (23.4, 1.0),
            "Feature16 (L17.2)": (17.2, 1.0),
            "Feature17 (2C)": (0.0, 2.0)
        }

    def read_di0_updates(self):
        """
        DI0 sinyallerini sürekli olarak stdin'den okur ve current_di0_value değerini günceller.
        Bu method bir thread içinde çalışır.
        """
        for line in sys.stdin:
            # print("")  # Log okunurluğu için boş satır
            try:
                # JSON formatına uygun olması için tekli tırnak işaretlerini çiftli tırnak işaretleriyle değiştir
                clean_line = line.strip().replace("'", '"')
                # print(f"Processing line: {clean_line}")
                data = json.loads(clean_line)
                if 'DI0' in data:
                    # JSON dizisini tuple'a çevir
                    di0_array = data['DI0']
                    self.current_di0_value = (di0_array[0], di0_array[1])
                    self.mech_eye.current_di0_value = self.current_di0_value
                    # print(f"Received DI0 update: {self.current_di0_value}")
            except json.JSONDecodeError as e:
                print(f"Error parsing input: {line}")
                print(f"JSON error: {str(e)}")
            except Exception as e:
                print(f"Error processing input: {e}")

    @staticmethod
    def rotate_point_cloud(points: np.ndarray, angle_degrees: float, axis: str) -> np.ndarray:
        angle_radians = np.radians(angle_degrees)
        if axis == "z":
            rotation_matrix = np.array([
                [np.cos(angle_radians), -np.sin(angle_radians), 0],
                [np.sin(angle_radians),  np.cos(angle_radians), 0],
                [0, 0, 1]
            ])
        elif axis == "x":
            rotation_matrix = np.array([
                [1, 0, 0],
                [0, np.cos(angle_radians), -np.sin(angle_radians)],
                [0, np.sin(angle_radians),  np.cos(angle_radians)]
            ])
        elif axis == "y":
            rotation_matrix = np.array([
                [np.cos(angle_radians), 0, np.sin(angle_radians)],
                [0, 1, 0],
                [-np.sin(angle_radians), 0, np.cos(angle_radians)]
            ])
        else:
            raise ValueError("Desteklenmeyen eksen. 'x', 'y' veya 'z' girin.")
        return points @ rotation_matrix.T

    @staticmethod
    def to_origin(points: np.ndarray) -> np.ndarray:
        min_x, min_y = np.min(points[:, 0]), np.min(points[:, 1])
        points[:, 0] -= min_x
        points[:, 1] -= min_y
        return points

    @staticmethod
    def remove_gripper_points(points: np.ndarray) -> np.ndarray:
        min_x = np.min(points[:, 0])
        y_candidates = points[:, 1][points[:, 0] < min_x + 23]
        y_min, y_max = np.min(y_candidates), np.max(y_candidates)
        return points[(points[:, 1] < y_min) | (points[:, 1] > y_max)]

    def get_next_valid_index(self, current_index: int, ignored: list, total_points: int) -> int:
        next_index = (current_index + 1) % total_points
        while ignored and next_index in ignored:
            next_index = (next_index + 1) % total_points
        return next_index

    def read_current_point_index(self) -> int:
        if os.path.exists(self.config.file_path):
            with open(self.config.file_path, 'r') as file:
                return int(file.read().strip())
        return 0

    def write_current_point_index(self, index: int):
        with open(self.config.file_path, 'w') as file:
            file.write(str(index))

    def pick_object(self, point: dict, soft_point, use_back_transit: bool):
        """
        Parça alma (pick) işlemini gerçekleştirir.
        
        Args:
            point (dict): Hedef nokta.
            soft_point: Soft nokta konumu.
            use_back_transit (bool): Back transit kullanılacak mı.
        """
        transit_vel = 80
        # Özel durum: Belirli noktalarda ekstra transit hareketleri
        if point == right_of_robot_points[0]:
            transit_vel = 50
            robot.MoveCart(p90, 0, 0, vel=self.config.vel_mul * transit_vel)
        elif point == left_of_robot_points[0]:
            transit_vel = 50
            robot.MoveCart(p91, 0, 0, vel=self.config.vel_mul * transit_vel)

        robot.MoveCart(soft_point, 0, 0, vel=self.config.vel_mul * 100)
        # Transit hareketleri
        if point in right_of_robot_points and point["p_up"][0] > 400:
            transit_vel = 50
            robot.MoveCart(right_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
        elif use_back_transit or (point["p_up"][0] > 400 and point in left_of_robot_points):
            transit_vel = 50
            robot.MoveCart(back_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
        
        # Parça alma hareketleri
        robot.MoveCart(point["p_up"], 0, 0, vel=self.config.vel_mul * transit_vel)
        robot.MoveL(point["p"], 0, 0, vel=self.config.vel_mul * 50)
        robot.WaitMs(1000)
        robot.SetDO(7, 1)  # Parçayı kavrama sinyali
        robot.WaitMs(1000)
        robot.MoveL(point["p_up"], 0, 0, vel=self.config.vel_mul * transit_vel)
        print("PICK OBJECT", self.current_di0_value)
        
        # DI0 değerine göre parça alma işleminin başarıyla tamamlanıp tamamlanmadığını kontrol et
        if not self.current_di0_value[1]:
            print("Parça başarıyla kavrandı.")
            robot.Mode(0)
        else:
            # Parça kavrama hatası durumunda
            robot.SetDO(7, 0)
            print("Parça kavurma hatası.")
            robot.Mode(1)
            
            # Yeni parça alma girişimi için indeksi güncelle
            current_index = self.read_current_point_index()
            current_index = self.get_next_valid_index(current_index, self.config.ignored_points, len(self.points))
            self.write_current_point_index(current_index)
            
            # Yeni parça alma işlemi için hazırlan
            point = self.points[current_index]
            if current_index < len(left_of_robot_points + left_small + back_points):
                soft_point = p90
            else:
                soft_point = p91
                
            total_left_back_points = len(left_of_robot_points + left_small + back_points)
            use_back_transit = current_index >= len(left_of_robot_points + left_small) and current_index < total_left_back_points
            
            if isinstance(self.config.same_place_index, int):
                point = self.points[self.config.same_place_index]
            else:
                point = self.points[current_index]
                
            self.pick_point = point
            self.pick_soft_point = soft_point
            self.pick_use_back_transit = use_back_transit
            self.pick_object(point, soft_point, use_back_transit)
            return
        
        # Geri dönüş transit hareketleri
        if point in right_of_robot_points and point["p_up"][0] > 400:
            transit_vel = 50
            robot.MoveCart(right_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
        elif use_back_transit or (point["p_up"][0] > 400 and (point in left_of_robot_points)):
            transit_vel = 50
            robot.MoveCart(back_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
        
        # Son hareket: Soft point'e git
        robot.MoveCart(soft_point, 0, 0, vel=self.config.vel_mul * transit_vel)

    def calculate_tolerance_distance(self, value: float, target: float, tolerance: float) -> float:
        return np.abs(value - target)

    def get_gradient_color(self, distance: float, tolerance: float) -> str:
        if distance > tolerance:
            return "FF0000"
        ratio = distance / tolerance
        red = int(255 * (1 - ratio))
        green = int(255 * (1 - ratio))
        blue = 255
        return f"{red:02X}{green:02X}{blue:02X}"

    def smol_calc(self, small_data: np.ndarray):
        small = self.rotate_point_cloud(small_data, -90, "z")
        small = small[small[:, 2] > np.min(small[:, 2]) + 37]
        small = small[small[:, 0] < np.min(small[:, 0]) + 50]
        small = self.to_origin(small)

        if self.config.save_point_clouds:
            self.pcd.points = o3d.utility.Vector3dVector(small)
            o3d.io.write_point_cloud("/home/eypan/Documents/JaguarInterface/flask-react-app/small.ply", self.pcd)

        circle_fitter = CircleFitter(small)
        _, z_center_small, radius_small = circle_fitter.fit_circles_and_plot(
            find_second_circle=False, val_x=0.18, val_z=0.2, delta_z=25
        )
        s_datum = circle_fitter.get_datum()
        self.dist_3mm_s, _ = circle_fitter.get_distance(second_crc=False, z_distance_to_datum=23.1)
        self.feature_3 = z_center_small - s_datum
        self.radius_small = radius_small
        self.z_center_small = z_center_small

    def hor_calc(self, horizontal_data: np.ndarray, horizontal2_data: np.ndarray):
        horizontal2_data[:, 2] -= 70
        horizontal2_data[:, 0] -= 50

        if self.config.save_point_clouds:
            self.pcd.points = o3d.utility.Vector3dVector(horizontal2_data)
            o3d.io.write_point_cloud("/home/eypan/Documents/JaguarInterface/flask-react-app/horizontal2.ply", self.pcd)
            self.pcd.points = o3d.utility.Vector3dVector(horizontal_data)
            o3d.io.write_point_cloud("/home/eypan/Documents/JaguarInterface/flask-react-app/horizontal_pre.ply", self.pcd)

        diff = np.abs(np.max(horizontal_data[:, 0]) - np.min(horizontal2_data[:, 0]))
        datum_horizontal = np.max(horizontal_data[:, 0]) - diff

        y_values = np.linspace(np.min(horizontal_data[:, 1]), np.max(horizontal_data[:, 1]), num=100)
        line_points = np.array([[datum_horizontal, y, np.min(horizontal_data[:, 2])] for y in y_values])
        augmented_pc = np.vstack((horizontal_data, line_points))
        horizontal = self.rotate_point_cloud(augmented_pc, 90, "z")
        horizontal = self.to_origin(horizontal)
        self.horizontal = horizontal

        if self.config.save_point_clouds:
            self.pcd.points = o3d.utility.Vector3dVector(horizontal)
            o3d.io.write_point_cloud("/home/eypan/Documents/JaguarInterface/flask-react-app/horizontal_post.ply", self.pcd)

        self.circle_fitter = CircleFitter(horizontal)
        _, circle2 = self.circle_fitter.fit_circles_and_plot()
        self.dist_3mm_h = self.circle_fitter.get_distance()[0]
        self.l_40 = get_40(horizontal)
        self.height = np.max(self.horizontal[:, 1]) - self.circle_fitter.get_datum()
        self.feature_1 = circle2[1]
        self.feature_2 = circle2[2]

    def process_vertical_measurement(self, vertical_data: np.ndarray) -> dict:
        """
        Vertical verileri işleyip ölçüm sonuçlarını hesaplar.
        
        Args:
            vertical_data (np.ndarray): Vertical tarama verisi.
            
        Returns:
            dict: Hesaplama sonuçlarını içeren sözlük.
        """
        vertical = self.remove_gripper_points(vertical_data)
        vertical = self.rotate_point_cloud(vertical, 180, "z")

        if self.config.save_point_clouds:
            pcd = o3d.geometry.PointCloud()
            pcd.points = o3d.utility.Vector3dVector(vertical)
            o3d.io.write_point_cloud("/home/eypan/Documents/JaguarInterface/flask-react-app/vertical.ply", pcd)
        
        vertical_copy = self.to_origin(vertical.copy())
        l_17_2, ok_17_2 = horn_diff(vertical_copy)
        l_23_4, ok_23_4 = horn_diff(vertical_copy, 240, 280)

        current_index = self.get_next_valid_index(self.read_current_point_index(), self.config.ignored_points, len(self.points))
        self.write_current_point_index(current_index)
        
        # vertical_results hesaplama
        B = self.circle_fitter.get_B()
        b_trans_val = np.max(vertical[:, 1]) - np.max(self.horizontal[:, 2])
        b_vertical = B + b_trans_val
        _, _, r1, l_79_73 = slope(vertical, b_vertical)
        _, _, r2, _ = slope(vertical, y_divisor=0.11, crc_l=27)
        l_42 = np.max(vertical[:, 1]) - b_vertical
        l_248 = arm_horn_lengths(vertical, b_vertical)
        mean_3mm = np.mean([self.dist_3mm_h, self.dist_3mm_s])
        l_88_6 = self.feature_1 - self.feature_2
        l_81_5 = filter_and_visualize_projection_with_ply(self.horizontal)

        vertical_results = {
            "l_17_2": l_17_2,
            "l_23_4": l_23_4,
            "l_42": l_42,
            "l_79_73": l_79_73,
            "l_248": l_248,
            "r1": (r1 - self.feature_2),
            "r2": (r2 + self.feature_2),
            "feature_1": self.feature_1,
            "mean_3mm": mean_3mm,
            "l_88_6": l_88_6,
            "l_81_5": l_81_5,
            "ok_17_2": ok_17_2
        }

        # current_results hesaplama
        current_results = self.combine_results(vertical_results)

        if hasattr(self, "pick_point"):
            if self.config.put_back:
                point = self.pick_point
                soft_point = self.pick_soft_point
                use_back_transit = self.pick_use_back_transit
                transit_vel = 80

                robot.MoveCart(soft_point, 0, 0, vel=self.config.vel_mul * 100)
                if point in right_of_robot_points and point["p_up"][0] > 400:
                    robot.MoveCart(right_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
                elif use_back_transit or (point["p_up"][0] > 400 and (point in left_of_robot_points)):
                    transit_vel = 50
                    robot.MoveCart(back_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
                
                robot.MoveCart(point["p_up"], 0, 0, vel=self.config.vel_mul * transit_vel)
                robot.MoveL(point["p"], 0, 0, vel=self.config.vel_mul * 50)
                robot.WaitMs(500)
                robot.SetDO(7, 0)
                robot.WaitMs(500)
                robot.MoveL(point["p_up"], 0, 0, vel=self.config.vel_mul * transit_vel)
                
                if point in right_of_robot_points and point["p_up"][0] > 400:
                    robot.MoveCart(right_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
                elif use_back_transit or (point["p_up"][0] > 400 and (point in left_of_robot_points)):
                    transit_vel = 50
                    robot.MoveCart(back_transit_point, 0, 0, vel=self.config.vel_mul * transit_vel)
            else:
                if self.config.drop_object:
                    print("DROP OBJECT", self.check_part_quality(current_results))
                    drop_point = metal_detector if self.check_part_quality(current_results) else trash
                    print("DROP POINT", drop_point)
                    robot.MoveCart(drop_point, 0, 0, vel=self.config.vel_mul * 80)
                    robot.SetDO(7, 0)
                else:
                    print("Parça bırakma ayarı devre dışı.")
        
        if current_index == len(self.points) - 1:
            self.write_current_point_index(0)   

        return vertical_results

    def check_part_quality(self, results: dict) -> bool:
        for feature, target_tolerance in self.tolerances.items():
            target, tolerance = target_tolerance
            value = results.get(feature, None)
            if value is None:
                print(f"Parça kalite kontrolünden geçemedi: {feature} değeri hesaplanamadı")
                return False
            if not (target - tolerance <= value <= target + tolerance):
                print(f"Parça kalite kontrolünden geçemedi: {feature} = {value}, hedef = {target}±{tolerance}")
                return False
        print("Parça kalite kontrolünden başarıyla geçti.")
        return True

    def combine_results(self, vertical_results: dict) -> dict:
        # Değerleri güvenli bir şekilde almak için yardımcı fonksiyon
        def safe_get(dict_obj, key, default=None):
            value = dict_obj.get(key, default)
            # Eğer bir değer sözlük ise, string temsiline dönüştür
            if isinstance(value, dict):
                print(f"Uyarı: {key} için dictionary değeri algılandı: {value}")
                return str(value)
            return value
        
        results = {
            "Feature1 (102.1)": safe_get(vertical_results, "feature_1", self.feature_1),
            "Feature2 (25mm/2)": self.feature_2,
            "Feature3 (23.1)": self.feature_3,
            "Feature4 (25mm/2)": self.radius_small,
            "Feature5 (L40)": self.l_40,
            "Feature6 (L248)": safe_get(vertical_results, "l_248"),
            "Feature7 (L42)": safe_get(vertical_results, "l_42"),
            "Feature8 (L79.73)": safe_get(vertical_results, "l_79_73"),
            "Feature9 (R1-50)": safe_get(vertical_results, "r1"),
            "Feature10 (R2-35)": safe_get(vertical_results, "r2"),
            "Feature11 (3mm)": safe_get(vertical_results, "mean_3mm"),
            "Feature12 (88.6)": safe_get(vertical_results, "l_88_6"),
            "Feature13 (10.6)": self.feature_3 - self.radius_small,
            "Feature14 (81.5)": safe_get(vertical_results, "l_81_5"),
            "Feature15 (L23.4)": safe_get(vertical_results, "l_23_4"),
            "Feature16 (L17.2)": safe_get(vertical_results, "l_17_2"),
            "Feature17 (2C)": safe_get(vertical_results, "ok_17_2")
        }
        
        # Tüm değerlerin serileştirilebilir olduğundan emin olalım
        for key, value in results.items():
            try:
                # Test amaçlı json.dumps
                json.dumps(value)
            except (TypeError, OverflowError):
                print(f"Serileştirilemeyen değer bulundu: {key} = {value}")
                results[key] = str(value)
        
        return results

    def write_excel_to_db(self, excel_path: str):
        df = pd.read_excel(excel_path, sheet_name="ScanResults")
        conn = mysql.connector.connect(
            host="192.168.1.180",
            user="cobot_dbuser",
            password="um6vv$7*sJ@5Q*",
            database="cobot"
        )
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS scan_results (
            id INT AUTO_INCREMENT PRIMARY KEY,
            iteration INT,
            feature VARCHAR(255),
            value DOUBLE
        )
        """
        cursor.execute(create_table_query)
        conn.commit()

        insert_query = "INSERT INTO scan_results (iteration, feature, value) VALUES (%s, %s, %s)"
        for _, row in df.iterrows():
            try:
                value_float = float(row['Value'])
            except ValueError:
                value_float = None
            data = (int(row['Iteration']), row['Feature'], value_float)
            cursor.execute(insert_query, data)
        
        conn.commit()
        cursor.close()
        conn.close()

    def save_to_excel(self):
        rows = []
        for iteration, result in enumerate(self.results, start=1):
            iteration_ok = 1
            
            # İlk hatanın nedenini görmek için detaylı loglama yapalım
            # print(f"İterasyon {iteration} sonuçları:")
            # print(result)
            
            for feature, value in result.items():
                # Dictionary değerlerini tespit edip düzgün şekilde dönüştürelim
                if isinstance(value, dict):
                    # Dictionary değeri string formatına dönüştürülüyor 
                    # print(f"Dictionary değeri tespit edildi: {feature} = {value}")
                    value_str = json.dumps(value)  # JSON string formatına dönüştürme
                    rows.append({"Iteration": iteration, "Feature": feature, "Value": value_str})
                else:
                    try:
                        # Değerin serileştirilebilir olduğundan emin olalım
                        rows.append({"Iteration": iteration, "Feature": feature, "Value": value})
                        
                        # Tolerans kontrolü
                        if feature in self.tolerances:
                            target, tolerance = self.tolerances[feature]
                            if not (target - tolerance <= value <= target + tolerance):
                                iteration_ok = 0
                    except TypeError as e:
                        # Serileştirilemez değer bulundu
                        print(f"Serileştirilemez değer: {feature} = {value}, hata: {e}")
                        rows.append({"Iteration": iteration, "Feature": feature, "Value": str(value)})
            
            rows.append({"Iteration": iteration, "Feature": "OK", "Value": iteration_ok})
        
        # DataFrame oluştur
        results_df = pd.DataFrame(rows)
        output_file = "/home/eypan/Documents/JaguarInterface/flask-react-app/scan_results.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, index=False, sheet_name="ScanResults")
                worksheet = writer.sheets["ScanResults"]
                
                for row_index, row in enumerate(results_df.itertuples(index=False), start=2):
                    if row.Feature == "OK":
                        row_fill = PatternFill(start_color="00B050" if row.Value else "FF0000", fill_type="solid")
                        worksheet.cell(row=row_index, column=3).fill = row_fill
                    else:        
                        iteration = row.Iteration
                        feature = row.Feature
                        value = row.Value
                        
                        # Satır arkaplan rengi
                        iteration_color = "FCE4D6" if iteration % 2 == 0 else "D9EAD3"
                        fill = PatternFill(start_color=iteration_color, end_color=iteration_color, fill_type="solid")
                        for col_index in range(1, len(results_df.columns) + 1):
                            worksheet.cell(row=row_index, column=col_index).fill = fill
                        
                        # Tolerans kontrolü sadece değer sayısal ise yapılmalı
                        if feature in self.tolerances:
                            try:
                                value_num = float(value) if isinstance(value, str) else value
                                target, tolerance = self.tolerances[feature]
                                
                                value_cell = worksheet.cell(row=row_index, column=3)
                                if target - tolerance <= value_num <= target + tolerance:
                                    value_cell.fill = PatternFill(start_color="00B050", fill_type="solid")
                                else:
                                    value_cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                                
                                distance = self.calculate_tolerance_distance(value_num, target, tolerance)
                                gradient_color = self.get_gradient_color(distance, tolerance)
                                distance_cell = worksheet.cell(row=row_index, column=4)
                                distance_cell.value = tolerance - distance
                                distance_cell.fill = PatternFill(start_color=gradient_color, fill_type="solid")
                            except (ValueError, TypeError):
                                print(f"Sayısal karşılaştırma yapılamadı: {feature} = {value}")
                                
            if self.config.save_to_db:
                self.write_excel_to_db(output_file)
                
        except Exception as e:
            print(f"Excel yazma hatası: {e}")
            import traceback
            traceback.print_exc()

    def run_scan_cycle(self):
        ROBOT_POSITIONS = self.config.robot_positions
        for self.cycle in range(self.config.range_):
            start_time = time.time()
            current_results = {}
            current_index = self.read_current_point_index()
            
            try:
                send_command({"cmd": 107, "data": {"content": "ResetAllError()"}})

                if self.config.pick:
                    if current_index < len(left_of_robot_points + left_small + back_points):
                        soft_point = p90
                    else:
                        soft_point = p91

                    total_left_back_points = len(left_of_robot_points + left_small + back_points)
                    use_back_transit = current_index >= len(left_of_robot_points + left_small) and current_index < total_left_back_points

                    if isinstance(self.config.same_place_index, int):
                        point = self.points[self.config.same_place_index]
                    else:
                        point = self.points[current_index]

                    self.pick_point = point
                    self.pick_soft_point = soft_point
                    self.pick_use_back_transit = use_back_transit

                    self.pick_object(point, soft_point, use_back_transit)

                robot.MoveCart(ROBOT_POSITIONS['scrc'], 0, 0, vel=self.config.vel_mul * 100)

                small_data = self.mech_eye.main(lua_name="small.lua", scan_line_count=1500)
                if self.config.use_agg:
                    thread_small = threading.Thread(target=self.smol_calc, args=(small_data,))
                    thread_small.start()
                else:
                    self.smol_calc(small_data)
                    plt.show()

                horizontal_data = self.mech_eye.main(lua_name="horizontal.lua", scan_line_count=1500)
                horizontal2_data = self.mech_eye.main(lua_name="horizontal2.lua", scan_line_count=1500)
                if self.config.use_agg:
                    thread_horizontal = threading.Thread(target=self.hor_calc, args=(horizontal_data, horizontal2_data))
                    thread_horizontal.start()
                else:
                    self.hor_calc(horizontal_data, horizontal2_data)
                    plt.show()

                vertical_data = self.mech_eye.main(lua_name="vertical.lua", scan_line_count=2500)
                plt.show()

                if self.config.use_agg:
                    thread_small.join()
                    thread_horizontal.join()

                vertical_results = self.process_vertical_measurement(vertical_data)
                current_results = self.combine_results(vertical_results)
                current_results["Processing Time (s)"] = time.time() - start_time
            except Exception as e:
                current_results = {"Error": str(e)}
            finally:
                self.results.append(current_results)
                if self.config.use_agg:
                    t_excel = threading.Thread(target=self.save_to_excel)
                    t_excel.start()
                    self.excel_threads.append(t_excel)
                else:
                    self.save_to_excel()

if __name__ == "__main__":
    scanner = JaguarScanner()
    scanner.run_scan_cycle()